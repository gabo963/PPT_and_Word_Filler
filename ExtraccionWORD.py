import docx
import re
import openpyxl

def sacarTexto( archivo ):
    
    documento = docx.Document( archivo )
    texto = []    
    
    for line in documento.paragraphs:
        
        encontrados = re.findall('\[[^\]]*\]', line.text)
            
        for encontrado in encontrados:
            texto.append( encontrado )
    
    texto = list(dict.fromkeys(texto))
    orden = []
    
    for linea in texto:
        orden.append(int(str(linea)[1:-1]))
    
    orden.sort() 
    
    texto = []
    
    for linea in orden:
        texto.append( "[" + str(linea) + "]" )
         
    return texto
    

def agregar_a_excelW( rutaWord, rutaExcel, me ):
    
    lista = sacarTexto( rutaWord )
    book = openpyxl.load_workbook( rutaExcel )
    
    if not 'Encontrar Parámetros Word' in book.sheetnames:
        book.create_sheet('Encontrar Parámetros Word')
    
    sheet = book['Encontrar Parámetros Word']
    
    fila = 3
    columna = 2
    
    i = 0
    aborrar = 0
    
    while sheet.cell( row=(i+fila), column=(columna) ).value != None:
        i += 1
    
    aborrar = i
    
    for i in range( aborrar ):
        sheet.cell( row=(i+fila), column=(columna) ).value = ""
        sheet.cell( row=(i+fila), column=(columna+1) ).value = ""
        sheet.cell( row=(i+fila), column=(columna+2) ).value = ""
        
    fila = 3
    columna = 2
    
    sheet.cell( row=(2), column=(2) ).value = 'Parámetro:'
    sheet.cell( row=(2), column=(3) ).value = 'Descripción:'
    sheet.cell( row=(2), column=(4) ).value = 'Valor:'
    
    for i in range( len(lista) ):
        sheet.cell( row=(i+fila), column=(columna) ).value = lista[i]
    
    book.save(rutaExcel)
    
    me['text'] = 'Hecho'
    me['state'] = 'disabled'
    me['command'] = None
